import io
from datetime import date

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

# ─────────────────────────────────────────────
# CONSTANTES DE ESTRUTURA DA PLANILHA
# ─────────────────────────────────────────────
IDX_QTDE      = 1   # pandas 0-based: linha "Qtde Posições"
IDX_DIA       = 2   # linha "Dia da Semana"
IDX_SEMANA    = 3   # linha "Semana"
IDX_DATA      = 4   # linha "Data"
IDX_GRUPO     = 5   # linha de grupo (MZ1, MZ2...)
IDX_POS_START = 6   # primeira linha de posições (0-based)

# Cores RGB (ARGB) consideradas verdes no Excel
GREEN_RGB = {"FF00FF00", "FF00B050", "FF92D050", "FF70AD47"}
YELLOW_RGB = {"FFFFFF00", "FFFFC000", "FFFFFF99", "FFFFEB9C"}
RED_RGB = {"FFFF0000", "FFC00000", "FFFF4444", "FFFF0000"}


def safe_rgb(fg):
    """Retorna o RGB da cor ou None sem lançar exceção."""
    try:
        if fg.type == "rgb":
            return str(fg.rgb).upper()
    except Exception:
        pass
    return None


def safe_theme(fg):
    """Retorna (theme, tint) ou (None, None) sem lançar exceção."""
    try:
        if fg.type == "theme":
            return fg.theme, (fg.tint or 0.0)
    except Exception:
        pass
    return None, None


def get_cell_status(cell):
    """Classifica a célula por cor. Nunca lança exceção."""
    try:
        fill = cell.fill
        if not fill or fill.fill_type != "solid":
            return "PENDENTE"
        fg = fill.fgColor
        if fg is None:
            return "PENDENTE"

        rgb = safe_rgb(fg)
        if rgb:
            if rgb == "00000000":
                return "PENDENTE"
            if rgb in GREEN_RGB:
                return "INVENTARIADO"
            if rgb in YELLOW_RGB:
                return "EM ANDAMENTO"
            if rgb in RED_RGB:
                return "PROBLEMA"
            return "PENDENTE"

        theme, tint = safe_theme(fg)
        if theme is not None:
            if theme == 9 and tint <= 0:   # verde escuro
                return "INVENTARIADO"
            if theme == 6 and tint <= 0:   # verde claro
                return "INVENTARIADO"
            if theme == 7:
                return "EM ANDAMENTO"
            if theme == 2:
                return "PROBLEMA"
        return "PENDENTE"
    except Exception:
        return "PENDENTE"


def count_green_cells(ws, col_idx, first_row, last_row):
    """Conta células verdes/amarelas/vermelhas em uma coluna, linha a linha."""
    inv = em_and = prob = 0
    for r in range(first_row, last_row + 1):
        status = get_cell_status(ws.cell(row=r, column=col_idx))
        if status == "INVENTARIADO":
            inv += 1
        elif status == "EM ANDAMENTO":
            em_and += 1
        elif status == "PROBLEMA":
            prob += 1
    return inv, em_and, prob


@st.cache_data(show_spinner=False)
def load_data(file_bytes, sheet_name):
    """
    Carrega e processa a planilha.
    Usa pandas para contar posições (rápido) e
    openpyxl apenas para ler cores (somente colunas com posições).
    """
    # ── 1. Pandas: estrutura e contagem de posições ──────────────────────────
    df_raw = pd.read_excel(
        io.BytesIO(file_bytes),
        sheet_name=sheet_name,
        header=None,
        dtype=str,          # tudo como string para evitar conversões
    )

    row_dia   = df_raw.iloc[IDX_DIA]
    row_data  = df_raw.iloc[IDX_DATA]
    posicoes  = df_raw.iloc[IDX_POS_START:]   # linhas de posições

    # Conta posições não-nulas por coluna (equivale ao CONT.VALORES)
    total_por_col = posicoes.apply(
        lambda col: col.dropna().apply(lambda v: str(v).strip() != "").sum()
    )

    # Filtra colunas válidas: tem data + é dia útil
    dias_uteis = ~row_dia.str.upper().isin(["SÁB", "SAB", "DOM"])
    tem_data   = row_data.notna() & (row_data != "Data")
    colunas_validas = df_raw.columns[dias_uteis & tem_data & (total_por_col > 0)]

    # Normaliza datas
    datas_raw = row_data[colunas_validas]
    datas = pd.to_datetime(datas_raw, errors="coerce")

    # ── 2. openpyxl: leitura de cores ────────────────────────────────────────
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    ws = wb[sheet_name]
    last_row = IDX_POS_START + len(posicoes)   # linha Excel (1-based) = pandas idx + 1

    registros = []
    for pandas_col_idx in colunas_validas:
        excel_col = pandas_col_idx + 1          # openpyxl é 1-based
        total = int(total_por_col[pandas_col_idx])
        if total == 0:
            continue

        inv, em_and, prob = count_green_cells(
            ws,
            col_idx=excel_col,
            first_row=IDX_POS_START + 1,        # +1 porque openpyxl é 1-based
            last_row=last_row,
        )
        pend = total - inv - em_and - prob

        registros.append({
            "Data":        datas[pandas_col_idx],
            "Dia":         str(row_dia[pandas_col_idx]),
            "Total":       total,
            "Inventariado": inv,
            "Em Andamento": em_and,
            "Problema":    prob,
            "Pendente":    max(pend, 0),
        })

    df = pd.DataFrame(registros)
    df = df.sort_values("Data").reset_index(drop=True)
    return df


# ─────────────────────────────────────────────
# STREAMLIT APP
# ─────────────────────────────────────────────

def main():
    st.set_page_config(page_title="Painel Micro Inventário", layout="wide")
    st.title("📦 Painel de Micro Inventário — CRONOGRAMA 2025")

    uploaded_file = st.file_uploader("📂 Envie o arquivo Excel do cronograma", type=["xlsx"])
    if not uploaded_file:
        st.info("Aguardando upload da planilha...")
        return

    st.sidebar.header("⚙️ Parâmetros")
    sheet_name = st.sidebar.text_input("Nome da aba", value="CRONOGRAMA")
    ignorar_passado = st.sidebar.checkbox(
        "Considerar apenas dias a partir de hoje nos cálculos de meta",
        value=False,
    )

    file_bytes = uploaded_file.read()

    with st.spinner("Processando planilha..."):
        try:
            df = load_data(file_bytes, sheet_name)
        except Exception as e:
            st.error(f"Erro ao processar o arquivo: {e}")
            return

    if df.empty:
        st.warning("Nenhuma posição encontrada. Verifique o nome da aba.")
        return

    today = pd.Timestamp(date.today())

    # ── Totais ────────────────────────────────────────────────────────────────
    total_pos  = int(df["Total"].sum())
    total_inv  = int(df["Inventariado"].sum())
    total_ea   = int(df["Em Andamento"].sum())
    total_prob = int(df["Problema"].sum())
    total_pend = int(df["Pendente"].sum())
    total_falta = total_ea + total_prob + total_pend
    pct_inv = round(total_inv / total_pos * 100, 1) if total_pos > 0 else 0.0

    # ── Dias abertos ──────────────────────────────────────────────────────────
    mask_aberto = df["Inventariado"] < df["Total"]
    if ignorar_passado:
        mask_aberto = mask_aberto & (df["Data"] >= today)
    dias_abertos_df = df[mask_aberto]
    n_dias = len(dias_abertos_df)
    ideal  = int((total_falta + n_dias - 1) / n_dias) if n_dias > 0 else 0

    # ── Métricas ──────────────────────────────────────────────────────────────
    st.subheader("📊 Resumo Geral")
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Total de Posições",    f"{total_pos:,}".replace(",", "."))
    c2.metric("✅ Inventariadas",      f"{total_inv:,}".replace(",", "."), f"{pct_inv}%")
    c3.metric("⬜ Pendentes",          f"{total_pend:,}".replace(",", "."))
    c4.metric("🟡 Em Andamento",       f"{total_ea:,}".replace(",", "."))
    c5.metric("🔴 Com Problema",       f"{total_prob:,}".replace(",", "."))

    st.subheader("🎯 Planejamento de Uniformidade")
    p1, p2, p3 = st.columns(3)
    p1.metric("Posições a inventariar",  f"{total_falta:,}".replace(",", "."))
    p2.metric("Dias úteis em aberto",    n_dias)
    p3.metric("Meta ideal por dia",      f"{ideal} pos/dia")

    st.info(
        f"💡 Para concluir o inventário de forma uniforme, processe "
        f"**{ideal} posições/dia** ao longo dos **{n_dias} dias úteis** restantes."
    )

    # ── Tabela ────────────────────────────────────────────────────────────────
    st.subheader("📋 Detalhamento por Dia")
    df_disp = df.copy()
    df_disp["Data"]       = df_disp["Data"].dt.strftime("%d/%m/%Y")
    df_disp["% Concluído"] = df.apply(
        lambda x: f"{round(x['Inventariado']/x['Total']*100,1)}%" if x["Total"] > 0 else "0%",
        axis=1,
    )
    df_disp["Meta Ideal"] = df.apply(
        lambda x: ideal if x["Inventariado"] < x["Total"] else "—", axis=1
    )
    st.dataframe(
        df_disp[["Data","Dia","Total","Inventariado","Em Andamento",
                 "Problema","Pendente","% Concluído","Meta Ideal"]],
        use_container_width=True,
        hide_index=True,
    )

    # ── Gráficos ──────────────────────────────────────────────────────────────
    st.subheader("📈 Gráficos")
    tab1, tab2 = st.tabs(["Progresso por Dia", "Meta vs Pendente (Dias Abertos)"])

    with tab1:
        chart = df.set_index("Data")[["Inventariado","Pendente","Em Andamento","Problema"]]
        st.bar_chart(chart, color=["#2ecc71","#bdc3c7","#f1c40f","#e74c3c"])

    with tab2:
        if n_dias > 0:
            meta_df = dias_abertos_df.copy().set_index("Data")[["Pendente"]]
            meta_df["Meta Ideal"] = ideal
            st.bar_chart(meta_df, color=["#bdc3c7","#3498db"])
        else:
            st.success("🎉 Todos os dias já foram concluídos!")

    # ── Legenda ───────────────────────────────────────────────────────────────
    st.subheader("ℹ️ Legenda de Cores")
    st.markdown("""
    | Cor da célula | Status |
    |---|---|
    | 🟢 Verde | Inventariado |
    | 🟡 Amarelo | Em Andamento |
    | 🔴 Vermelho | Com Problema |
    | ⬜ Sem cor | Pendente |
    """)


if __name__ == "__main__":
    main()