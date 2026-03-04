import io
from datetime import date, datetime

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
import requests

# ─────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────
IDX_DATA_ROW  = 5   # Excel row: datas
IDX_GROUP_ROW = 6   # Excel row: grupos (MZ1, EST...)
IDX_POS_START = 7   # Excel row: início das posições

GREEN_RGB = "FF00FF00"

CACHE_VERSION = "2026-03-04-v3"  # altere este valor sempre que mudar o código de parsing

GITHUB_URL = (
    "https://github.com/Djalmandre/Inventario26/raw/refs/heads/main/"
    "CRONOGRAMA%202026%20RECAP.xlsm"
)


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def fetch_file_bytes(url: str) -> bytes:
    headers = {"User-Agent": "streamlit-app"}
    resp = requests.get(url, headers=headers, timeout=90)
    resp.raise_for_status()
    return resp.content


def parse_excel_date(value) -> pd.Timestamp | None:
    """
    Converte o valor bruto da célula de data para pd.Timestamp.
    - datetime/date  → converte direto
    - int/float      → serial Excel (dias desde 1899-12-30)
    - str            → tenta parse; descarta rótulos como 'Data'
    - Retorna None se inválido ou anterior a 2000-01-01
    """
    if value is None:
        return None

    ts = None

    if isinstance(value, (datetime, date)):
        ts = pd.Timestamp(value)

    elif isinstance(value, (int, float)):
        # Serial Excel: 1 = 1900-01-01
        try:
            ts = pd.Timestamp("1899-12-30") + pd.Timedelta(days=int(value))
        except Exception:
            return None

    elif isinstance(value, str):
        try:
            ts = pd.to_datetime(value, dayfirst=True)
        except Exception:
            return None

    if ts is None:
        return None

    # Descarta datas fora do intervalo esperado (evita seriais errados)
    if ts.year < 2000 or ts.year > 2100:
        return None

    return ts


# ─────────────────────────────────────────────
# CARGA DE DADOS
# ─────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def load_data(file_bytes: bytes, sheet_name: str, _cache_version: str = CACHE_VERSION) -> pd.DataFrame:
    """
    Lê a planilha em modo read_only.
    Conta posições ÚNICAS inventariadas — evita duplicatas
    quando a mesma posição está verde em mais de uma coluna/dia.
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)

    # Alteração 3: valida se a aba existe antes de acessar
    if sheet_name not in wb.sheetnames:
        raise KeyError(
            f"Aba '{sheet_name}' não encontrada. "
            f"Abas disponíveis: {wb.sheetnames}"
        )

    ws = wb[sheet_name]

    col_data       = {}   # col_idx -> pd.Timestamp
    col_grupo      = {}   # col_idx -> str
    col_total      = {}   # col_idx -> int
    col_verde_vals = {}   # col_idx -> set de valores únicos verdes

    for row in ws.iter_rows(min_row=1):
        row_num = None
        for cell in row:
            if hasattr(cell, "row") and cell.row is not None:
                row_num = cell.row
                break
        if row_num is None:
            continue

        # ── Linha de datas ────────────────────────────────────────────────
        if row_num == IDX_DATA_ROW:
            for cell in row:
                if not hasattr(cell, "column") or cell.value is None:
                    continue
                ts = parse_excel_date(cell.value)
                if ts is not None:
                    col_data[cell.column] = ts

        # ── Linha de grupos ───────────────────────────────────────────────
        elif row_num == IDX_GROUP_ROW:
            for cell in row:
                if hasattr(cell, "column") and cell.value is not None:
                    col_grupo[cell.column] = str(cell.value)

        # ── Linhas de posições ────────────────────────────────────────────
        elif row_num >= IDX_POS_START:
            for cell in row:
                if not hasattr(cell, "column"):
                    continue
                c = cell.column
                if c not in col_data:          # ignora colunas sem data válida
                    continue
                if cell.value is None or str(cell.value).strip() == "":
                    continue

                col_total[c] = col_total.get(c, 0) + 1

                try:
                    fill = cell.fill
                    if (fill and fill.fill_type == "solid"
                            and fill.fgColor
                            and fill.fgColor.type == "rgb"
                            and str(fill.fgColor.rgb).upper() == GREEN_RGB):
                        val = str(cell.value).strip().upper()
                        col_verde_vals.setdefault(c, set()).add(val)
                except Exception:
                    pass

    wb.close()

    # Cada posição contada UMA única vez (primeira ocorrência verde)
    ja_inventariadas: set = set()
    records = []

    for c in sorted(col_data.keys()):
        total = col_total.get(c, 0)
        if total == 0:
            continue

        verdes_col = col_verde_vals.get(c, set())
        novas      = verdes_col - ja_inventariadas
        ja_inventariadas |= verdes_col

        records.append({
            "col":          c,
            "Data":         col_data[c],
            "Grupo":        col_grupo.get(c, ""),
            "Total":        total,
            "Inventariado": len(novas),
            "Pendente":     total - len(novas),
        })

    df = pd.DataFrame(records).sort_values("Data").reset_index(drop=True)
    return df


# ─────────────────────────────────────────────
# APP PRINCIPAL
# ─────────────────────────────────────────────
def main():
    st.set_page_config(page_title="Painel Micro Inventário", layout="wide")
    st.title("📦 Painel de Micro Inventário — CRONOGRAMA 2026")

    # ── Sidebar ───────────────────────────────────────────────────────────
    st.sidebar.header("⚙️ Parâmetros")
    sheet_name = st.sidebar.text_input("Nome da aba", value="CRONOGRAMA")
    ignorar_passado = st.sidebar.checkbox(
        "Considerar apenas dias a partir de hoje nos cálculos de meta",
        value=False,
    )
    st.sidebar.caption("📂 Planilha carregada automaticamente do GitHub.")

    # ── Download ──────────────────────────────────────────────────────────
    with st.spinner("⬇️ Baixando planilha do GitHub..."):
        try:
            file_bytes = fetch_file_bytes(GITHUB_URL)
        except Exception as e:
            st.error(f"Erro ao baixar arquivo: {e}")
            st.stop()

    # ── Processamento ─────────────────────────────────────────────────────
    with st.spinner("🔍 Lendo células e cores da planilha..."):
        try:
            df = load_data(file_bytes, sheet_name)
        except Exception as e:
            st.error(f"Erro ao processar planilha: {e}")
            st.stop()

    # Alteração 2: validações explícitas do DataFrame
    if df is None or df.empty:
        st.error("load_data retornou vazio. Verifique se a aba existe e se a linha de datas está correta.")
        st.stop()

    if "Data" not in df.columns:
        st.error(f"Coluna 'Data' não foi criada. Colunas retornadas: {list(df.columns)}")
        st.stop()

    if df["Data"].isna().all():
        st.error("Coluna 'Data' existe, mas todas as datas estão inválidas (NaT). Verifique a linha 5 da planilha.")
        st.stop()

    today = pd.Timestamp(date.today())

    # ── Totais globais ────────────────────────────────────────────────────
    total_pos  = int(df["Total"].sum())
    total_inv  = int(df["Inventariado"].sum())
    total_pend = int(df["Pendente"].sum())
    pct_inv    = round(total_inv / total_pos * 100, 1) if total_pos > 0 else 0.0

    mask_aberto = df["Pendente"] > 0
    if ignorar_passado:
        mask_aberto = mask_aberto & (df["Data"] >= today)
    dias_abertos_df = df[mask_aberto]
    n_dias = len(dias_abertos_df)
    ideal  = int((total_pend + n_dias - 1) / n_dias) if n_dias > 0 else 0

    # ── Métricas ──────────────────────────────────────────────────────────
    st.subheader("📊 Resumo Geral")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("📦 Total de Posições",  f"{total_pos:,}".replace(",", "."))
    c2.metric("✅ Inventariadas",       f"{total_inv:,}".replace(",", "."), f"{pct_inv}%")
    c3.metric("⬜ Pendentes",           f"{total_pend:,}".replace(",", "."))
    c4.metric("📅 Dias úteis em aberto", n_dias)

    # ── Planejamento ──────────────────────────────────────────────────────
    st.subheader("🎯 Planejamento de Uniformidade")
    p1, p2 = st.columns(2)
    p1.metric("Meta ideal por dia", f"{ideal} pos/dia")
    p2.metric("Progresso geral",    f"{pct_inv}%")

    st.info(
        f"💡 Para concluir o inventário de forma uniforme, processe "
        f"**{ideal} posições/dia** ao longo de **{n_dias} dias úteis**."
    )

    # ── Tabela detalhada ──────────────────────────────────────────────────
    st.subheader("📋 Detalhamento por Dia")
    df_disp = df.copy()
    df_disp["Data"] = df_disp["Data"].dt.strftime("%d/%m/%Y")
    df_disp["% Concluído"] = df.apply(
        lambda x: f"{round(x['Inventariado'] / x['Total'] * 100, 1)}%"
        if x["Total"] > 0 else "0%",
        axis=1,
    )
    df_disp["Meta Ideal"] = df.apply(
        lambda x: f"{ideal} pos" if x["Pendente"] > 0 else "✅ Concluído",
        axis=1,
    )
    st.dataframe(
        df_disp[["Data", "Grupo", "Total", "Inventariado", "Pendente",
                 "% Concluído", "Meta Ideal"]],
        use_container_width=True,
        hide_index=True,
    )

    # ── Gráficos ──────────────────────────────────────────────────────────
    st.subheader("📈 Gráficos")
    tab1, tab2, tab3 = st.tabs([
        "📊 Progresso por Dia",
        "🎯 Meta vs Pendente",
        "📉 Curva de Progresso Acumulado",
    ])

    with tab1:
        chart_df = df.copy()
        chart_df["Data_str"] = chart_df["Data"].dt.strftime("%d/%m")
        chart_df = chart_df.set_index("Data_str")[["Inventariado", "Pendente"]]
        st.bar_chart(chart_df, color=["#2ecc71", "#bdc3c7"])

    with tab2:
        if n_dias > 0:
            meta_df = dias_abertos_df.copy()
            meta_df["Data_str"] = meta_df["Data"].dt.strftime("%d/%m")
            meta_df = meta_df.set_index("Data_str")[["Pendente"]]
            meta_df["Meta Ideal"] = ideal
            st.bar_chart(meta_df, color=["#bdc3c7", "#3498db"])
        else:
            st.success("🎉 Todos os dias já foram concluídos!")

    with tab3:
        # Curva acumulada de inventário
        df_sorted = df.sort_values("Data").copy()
        df_sorted["Inv. Acumulado"] = df_sorted["Inventariado"].cumsum()
        df_sorted["Total Acumulado"] = df_sorted["Total"].cumsum()
        df_sorted["Data_str"] = df_sorted["Data"].dt.strftime("%d/%m")
        curva_df = df_sorted.set_index("Data_str")[["Inv. Acumulado", "Total Acumulado"]]
        st.line_chart(curva_df, color=["#2ecc71", "#3498db"])
        st.caption("🟢 Inventariado acumulado  |  🔵 Total de posições acumulado")

    # ── Legenda ───────────────────────────────────────────────────────────
    st.subheader("ℹ️ Legenda")
    st.markdown("""
    | Cor da célula | Status |
    |---|---|
    | 🟢 Verde (`FF00FF00`) | Inventariado |
    | ⬜ Sem cor | Pendente |

    > **Nota:** Cada posição é contada **uma única vez**, mesmo que apareça verde em múltiplos dias.
    """)


if __name__ == "__main__":
    main()
